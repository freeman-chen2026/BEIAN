import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from copy import copy

st.set_page_config(page_title="每日通航运行情况跟踪表生成器", layout="wide")
st.title("🛫 每日通航运行情况跟踪表生成器")
st.markdown("上传模板Excel和数据Excel，生成带格式的备案表格。")

# ---------- 注册号 -> ICAO 机型映射 ----------
DEFAULT_ICAO_MAP = {
    "B65AP": "GLF4",
    "B652R": "GLF4",
    "B652S": "GLF4",
    "B8105": "GLEX",
    "B8309": "GLF5",
    "B652Q": "GLF4",
    "B3926": "LJ60",
    "B8160": "GLF5",
    "B8262": "GLF4",
    "B8292": "GLF5",
}

# ---------- 辅助函数 ----------
def find_header_row(df_raw):
    keywords = ["客户", "航班号", "飞机注册号", "出发日期"]
    for idx, row in df_raw.iterrows():
        row_str = " ".join([str(v) for v in row.values if pd.notna(v)])
        if all(kw in row_str for kw in keywords):
            return idx
    return 0

def parse_uploaded_file(uploaded_file):
    df_all = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    header_row = find_header_row(df_all)
    df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row)
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    return df

def map_usage(usage):
    usage = str(usage).strip()
    if "调机" in usage:
        return "调机飞行", "调机飞行"
    elif "载客" in usage or "共享租赁" in usage:
        return "公务飞行", "私用飞行"
    else:
        return "公务飞行", "私用飞行"

def format_time(value):
    if pd.isna(value) or value == "" or value is None:
        return ""
    if isinstance(value, str):
        if ":" in value:
            parts = value.split(":")
            if len(parts) == 2:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:00"
            elif len(parts) == 3:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:{parts[2].zfill(2)}"
        return value
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M:%S")
    return str(value)

# ---------- 侧边栏：映射管理 ----------
st.sidebar.header("✈️ 注册号 → ICAO 机型映射")
user_mapping_text = st.sidebar.text_area(
    "自定义映射（覆盖内置）",
    value="\n".join([f"{k} {v}" for k, v in DEFAULT_ICAO_MAP.items()]),
    height=150
)
def parse_mapping(text):
    map_dict = {}
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split()
        if len(parts) >= 2:
            reg, icao = parts[0], parts[1]
            map_dict[reg.upper()] = icao.upper()
    return map_dict
icao_map = parse_mapping(user_mapping_text)
for k, v in DEFAULT_ICAO_MAP.items():
    if k not in icao_map:
        icao_map[k] = v

# ---------- 主界面 ----------
st.subheader("📂 上传文件")
template_file = st.file_uploader("【模板】上传带格式的 Excel 模板（如 每日通航运行情况跟踪表（天成商务航空有限公司）20260708.xlsx）", type=["xlsx"], key="template")
data_file = st.file_uploader("【数据】上传航段数据导出 Excel（如 航段数据导出 (40).xlsx）", type=["xlsx"], key="data")

if template_file and data_file:
    try:
        # 1. 读取数据
        df_raw = parse_uploaded_file(data_file)
        st.success(f"✅ 成功读取 {len(df_raw)} 条航段记录")

        # 2. 加载模板工作簿
        template_io = BytesIO(template_file.read())
        wb = load_workbook(template_io)
        ws = wb.active

        # 3. 确定表头行和数据起始行（假设表头在第2行，数据从第3行开始）
        header_row = 2
        data_start_row = header_row + 1

        # 4. 获取数据起始行（第3行）的样式，用于新增行
        style_row = data_start_row
        style_cells = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=style_row, column=col)
            style_cells[col] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format
            }

        # 5. 清空旧数据（从 data_start_row 到最大行）
        max_row = ws.max_row
        if max_row >= data_start_row:
            ws.delete_rows(data_start_row, max_row - data_start_row + 1)

        # 6. 准备新数据（按原始顺序）
        has_actual_depart = "实际出发" in df_raw.columns
        records = []
        for _, row in df_raw.iterrows():
            dt = pd.to_datetime(row["出发日期"])
            flight_date = f"{dt.year}/{dt.month}/{dt.day}"

            dep_city = str(row.get("出发城市", "")).strip()
            arr_city = str(row.get("到达城市", "")).strip()
            route = f"{dep_city}-{arr_city}" if dep_city and arr_city else f"{row['出发地']}-{row['到达地']}"
            run_type, oper_type = map_usage(row["用途"])

            if has_actual_depart:
                actual_depart = format_time(row.get("实际出发", ""))
                start_time = actual_depart if actual_depart else format_time(row.get("计划出发", ""))
            else:
                start_time = format_time(row.get("计划出发", ""))

            est_landing = format_time(row.get("预计到达", ""))
            actual_end = format_time(row.get("实际到达", "")) if "实际到达" in df_raw.columns else ""
            status = str(row.get("航段状态", "")).strip()
            is_landed = "是" if status in ["已执飞", "已完成"] else "否"
            reg = str(row["飞机注册号"]).strip().upper()
            icao_type = icao_map.get(reg, "")

            record = {
                "所属监管局": "深圳局",
                "运行人标准名称": "天成商务航空有限公司",
                "飞行活动的日期": flight_date,
                "当日飞行的运行种类": run_type,
                "当日飞行的经营种类": oper_type,
                "航空器型号": icao_type,
                "航空器注册号": reg,
                "是否向监控中心完成计划备案": "是",
                "是否获得飞行计划部门批准飞行": "是",
                "飞行开始时间": start_time,
                "飞行预计落地时间": est_landing,
                "是否已落地": is_landed,
                "飞行实际结束时间": actual_end if is_landed == "是" else "",
                "飞行地点（航线）": route,
                "选择允许的运行种类": "3.公务航空运行",
                "监管局是否电话跟踪该飞行动态": ""
            }
            records.append(record)

        # 7. 写入新数据（如果数据行数超过模板原有样式行数，新增行复制样式）
        for i, rec in enumerate(records):
            row_num = data_start_row + i
            # 如果当前行超出原有最大行，需要插入新行并复制样式
            if row_num > ws.max_row:
                ws.insert_rows(row_num)
                # 复制上一行的样式到新行
                for col in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=row_num - 1, column=col)
                    dst_cell = ws.cell(row=row_num, column=col)
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format

            col_names = [
                "所属监管局", "运行人标准名称", "飞行活动的日期", "当日飞行的运行种类", "当日飞行的经营种类",
                "航空器型号", "航空器注册号", "是否向监控中心完成计划备案", "是否获得飞行计划部门批准飞行",
                "飞行开始时间", "飞行预计落地时间", "是否已落地", "飞行实际结束时间", "飞行地点（航线）",
                "选择允许的运行种类", "监管局是否电话跟踪该飞行动态"
            ]
            for col_idx, col_name in enumerate(col_names, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = rec[col_name]
                # 确保样式正确（如果新增行已复制样式，这里无需再复制；但为了安全，也可以重新应用样式）
                # 对于已有行，我们也重新应用样式以保证格式统一
                # 但为了保留模板原有样式（如某些行可能特殊），我们只在新行复制样式。
                # 这里我们简单处理：如果该行是新增的（row_num > 原始行数），样式已复制，无需再处理。
                # 如果该行原本存在，我们保留其原有样式（因为只是覆盖值）
                # 所以这里只赋值，不改变样式。
                # 但为了统一，我们可以对所有行应用样式，但不覆盖可能不同的特殊格式（如合并单元格等）。
                # 最佳实践：只赋值，样式已在之前复制或保留。
                pass

        # 8. 如果数据行数少于模板原来行数，多余行已被删除（步骤5已删除全部旧数据）
        # 此时写入数据后，行数正好等于数据条数，若模板原来有多余空行，已被删除。

        # 9. 保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 预览
        preview_df = pd.DataFrame(records)
        st.subheader("📋 数据预览（按原始顺序）")
        st.dataframe(preview_df, use_container_width=True)

        st.download_button(
            label="⬇️ 下载带格式的 Excel 文件",
            data=output,
            file_name="每日通航运行情况跟踪表_生成.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ 处理失败：{e}")
        st.exception(e)
else:
    st.info("👆 请同时上传模板和数据文件。")
